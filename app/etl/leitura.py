"""Leitura segura de planilhas.

Nenhuma fórmula/macro é executada: openpyxl abre em `data_only` e o pandas
recebe apenas valores. Arquivos .xls e .csv também são aceitos.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl
import pandas as pd

from app.config import config
from app.utils.erros import ErroValidacaoArquivo
from app.utils.formato import numero
from app.utils.log import get_logger
from app.utils.texto import normalizar_coluna

logger = get_logger("etl.leitura")

MAX_LINHAS_CABECALHO = 8


@dataclass
class Planilha:
    nome: str
    dados: pd.DataFrame
    linha_cabecalho: int

    @property
    def colunas_normalizadas(self) -> list[str]:
        return [normalizar_coluna(c) for c in self.dados.columns]


def validar_arquivo(caminho: Path, tamanho_bytes: int | None = None) -> None:
    """Extensão, tamanho e nome — antes de qualquer leitura (regra 47)."""
    extensao = caminho.suffix.lower()
    if extensao not in config.EXTENSOES_PERMITIDAS:
        permitidas = ", ".join(sorted(config.EXTENSOES_PERMITIDAS))
        raise ErroValidacaoArquivo(
            f"Arquivo '{caminho.name}' não é aceito. Envie um arquivo {permitidas}."
        )
    tamanho = tamanho_bytes if tamanho_bytes is not None else caminho.stat().st_size
    limite = config.TAMANHO_MAXIMO_MB * 1024 * 1024
    if tamanho > limite:
        raise ErroValidacaoArquivo(
            f"Arquivo '{caminho.name}' tem {tamanho / 1024 / 1024:.1f} MB e o limite é "
            f"{config.TAMANHO_MAXIMO_MB} MB."
        )
    if tamanho == 0:
        raise ErroValidacaoArquivo(f"Arquivo '{caminho.name}' está vazio.")


def _detectar_cabecalho(bruto: pd.DataFrame) -> int:
    """Descobre em qual linha está o cabeçalho (planilhas com título/logo em cima)."""
    melhor_linha, melhor_pontos = 0, -1
    for indice in range(min(MAX_LINHAS_CABECALHO, len(bruto))):
        linha = bruto.iloc[indice]
        preenchidos = [v for v in linha if not pd.isna(v) and str(v).strip() != ""]
        textos = [v for v in preenchidos if isinstance(v, str)]
        pontos = len(preenchidos) + len(textos) * 2
        if len(preenchidos) < 2:
            continue
        if pontos > melhor_pontos:
            melhor_linha, melhor_pontos = indice, pontos
    return melhor_linha


def _mensagem_arquivo_grande(caminho: Path, total_linhas: int) -> ErroValidacaoArquivo:
    texto = (
        f"Arquivo '{caminho.name}' tem aproximadamente {numero(total_linhas)} linhas, acima "
        f"do limite de {numero(config.LIMITE_LINHAS_ARQUIVO)} linhas suportado por esta "
        "instância."
    )
    return ErroValidacaoArquivo(texto, [
        "Divida o arquivo em partes menores (ex.: por mês ou por cidade) e envie cada "
        "parte separadamente — reenviar não duplica registros, cada linha é identificada "
        "pela própria chave (data, cidade, equipe...).",
        "Ou, se esse volume for a rotina, migre para um plano com mais memória — "
        "veja docs/deploy.md.",
    ])


LIMITE_VARREDURA_FANTASMA = 500_000


def _contar_linhas_reais_xlsx(livro, limite: int) -> int:
    """Conta linhas com pelo menos uma célula preenchida, parando assim que
    tiver certeza — ou porque já passou do limite, ou porque já varreu
    `LIMITE_VARREDURA_FANTASMA` linhas cruas sem achar tantas preenchidas
    (arquivo com dimensão "fantasma", ver `estimar_linhas_xlsx`)."""
    total_real, varridas = 0, 0
    for aba in livro.worksheets:
        for linha in aba.iter_rows(values_only=True):
            varridas += 1
            if any(v is not None and str(v).strip() != "" for v in linha):
                total_real += 1
                if total_real > limite:
                    return total_real
            if varridas >= LIMITE_VARREDURA_FANTASMA:
                return total_real
    return total_real


def estimar_linhas_xlsx(caminho: Path, limite: int | None = None) -> int:
    """Estima o total de linhas com dado real, gastando o mínimo possível.

    `max_row` do openpyxl em modo `read_only` é O(1) (lê a dimensão
    declarada no XML, sem abrir células) — cobre a maioria dos arquivos
    instantaneamente. Mas planilhas exportadas por outros sistemas
    frequentemente têm formatação/estilo arrastados além dos dados reais
    (ex.: alguém formatou "a coluna inteira"), fazendo o openpyxl reportar
    um `max_row` muito maior do que as linhas realmente preenchidas — nesse
    caso o pandas, na prática, só materializa as linhas com conteúdo. Por
    isso, só quando `max_row` já aponta acima do limite é que confirmamos
    com uma varredura real (limitada) em vez de confiar cegamente nele —
    sem essa confirmação, um arquivo pequeno com esse tipo de formatação
    seria recusado por engano.
    """
    limite = config.LIMITE_LINHAS_ARQUIVO if limite is None else limite
    try:
        livro = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001 - deixa o pandas tentar e dar a mensagem de erro
        return 0
    try:
        max_row_declarado = sum(aba.max_row or 0 for aba in livro.worksheets)
        if max_row_declarado <= limite:
            return max_row_declarado
        return _contar_linhas_reais_xlsx(livro, limite)
    finally:
        livro.close()


def estimar_linhas_csv(caminho: Path) -> int:
    """Conta as linhas lendo em blocos, sem carregar o arquivo inteiro na
    memória de uma vez."""
    total_linhas = 0
    with caminho.open("rb") as arquivo:
        while bloco := arquivo.read(1024 * 1024):
            total_linhas += bloco.count(b"\n")
    return total_linhas


def estimar_linhas_arquivo(caminho: Path) -> int:
    """Estimativa cheap de linhas para qualquer extensão aceita — usada
    para somar o total de um LOTE de arquivos antes de processar (ver
    `app.routes.api.upload`). `.xls` legado não tem uma forma barata de
    contar sem ler tudo, mas o próprio formato trava em 65.536 linhas por
    aba, bem abaixo do limite — devolve 0 (não contribui para a soma).
    """
    extensao = caminho.suffix.lower()
    try:
        if extensao in (".xlsx", ".xlsm"):
            return estimar_linhas_xlsx(caminho)
        if extensao == ".csv":
            return estimar_linhas_csv(caminho)
    except Exception:  # noqa: BLE001 - estimativa é best-effort, nunca derruba o upload
        logger.warning("Não foi possível estimar linhas de %s", caminho.name, exc_info=True)
    return 0


def _validar_linhas_xlsx(caminho: Path) -> None:
    total_linhas = estimar_linhas_xlsx(caminho)
    if total_linhas > config.LIMITE_LINHAS_ARQUIVO:
        raise _mensagem_arquivo_grande(caminho, total_linhas)


def _validar_linhas_csv(caminho: Path) -> None:
    total_linhas = estimar_linhas_csv(caminho)
    if total_linhas > config.LIMITE_LINHAS_ARQUIVO:
        raise _mensagem_arquivo_grande(caminho, total_linhas)


def _limpar(df: pd.DataFrame) -> pd.DataFrame:
    """Remove colunas/linhas totalmente vazias e nomeia colunas sem cabeçalho.

    Propositalmente NÃO reindexa as linhas: o índice original (posição na
    aba, 0-based, igual ao de `pd.read_excel(..., header=None)`) é mantido
    para que o índice + 1 corresponda ao número real da linha no Excel —
    isso é o que permite apontar "linha 42" numa mensagem de erro.
    """
    df = df.dropna(axis=1, how="all").dropna(axis=0, how="all")
    df.columns = [
        str(c).strip() if not str(c).startswith("Unnamed") else f"coluna_{i}"
        for i, c in enumerate(df.columns)
    ]
    return df


def ler_planilhas(caminho: Path) -> list[Planilha]:
    """Devolve todas as abas úteis do arquivo."""
    validar_arquivo(caminho)
    extensao = caminho.suffix.lower()

    if extensao == ".csv":
        _validar_linhas_csv(caminho)
        for separador in (";", ",", "\t"):
            try:
                bruto = pd.read_csv(caminho, sep=separador, dtype=object,
                                    encoding="utf-8-sig", header=None)
            except (UnicodeDecodeError, pd.errors.ParserError):
                try:
                    bruto = pd.read_csv(caminho, sep=separador, dtype=object,
                                        encoding="latin-1", header=None)
                except Exception:  # noqa: BLE001 - tenta o próximo separador
                    continue
            if bruto.shape[1] > 1:
                linha = _detectar_cabecalho(bruto)
                dados = bruto.iloc[linha + 1:].copy()
                dados.columns = bruto.iloc[linha]
                return [Planilha(caminho.stem, _limpar(dados), linha)]
        raise ErroValidacaoArquivo(
            f"Não foi possível interpretar o CSV '{caminho.name}'. "
            "Salve-o novamente em Excel (.xlsx) e tente de novo."
        )

    motor = "openpyxl" if extensao in (".xlsx", ".xlsm") else None
    if motor == "openpyxl":
        _validar_linhas_xlsx(caminho)
    # read_only faz o openpyxl abrir o arquivo em modo streaming em vez de
    # montar a árvore completa de células em memória — sem isso, um
    # arquivo de dezenas de milhares de linhas podia dobrar o uso de
    # memória (a árvore do openpyxl + o DataFrame do pandas coexistindo) e
    # estourar o limite de RAM de instâncias pequenas (ex.: 512 MB no
    # plano gratuito do Render), derrubando o processo no meio do upload.
    argumentos_motor = {"read_only": True} if motor == "openpyxl" else {}
    try:
        abas = pd.read_excel(caminho, sheet_name=None, header=None, dtype=object, engine=motor,
                             engine_kwargs=argumentos_motor)
    except Exception as erro:  # noqa: BLE001 - convertido em mensagem amigável
        logger.exception("Falha ao abrir %s", caminho.name)
        raise ErroValidacaoArquivo(
            f"Não foi possível abrir '{caminho.name}'. Verifique se o arquivo não está "
            "corrompido nem protegido por senha.",
            [str(erro)[:200]],
        ) from erro

    planilhas: list[Planilha] = []
    for nome, bruto in abas.items():
        if bruto.empty or bruto.dropna(how="all").empty:
            continue
        linha = _detectar_cabecalho(bruto)
        dados = bruto.iloc[linha + 1:].copy()
        dados.columns = bruto.iloc[linha]
        dados = _limpar(dados)
        if dados.empty:
            continue
        planilhas.append(Planilha(str(nome), dados, linha))

    if not planilhas:
        raise ErroValidacaoArquivo(
            f"O arquivo '{caminho.name}' não possui nenhuma planilha com dados."
        )
    return planilhas
