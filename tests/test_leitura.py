"""Leitura de arquivos: contagem de linhas, dimensão fantasma e leitura em blocos.

Arquivo grande não é recusado — é lido em pedaços (`ler_em_blocos`), o que
mantém a memória constante. `ler_planilhas` (que carrega tudo de uma vez)
continua recusando, e é o pipeline que decide qual caminho usar.
"""
import openpyxl
import pandas as pd
import pytest

from app.config import config
from app.etl.leitura import (estimar_linhas_arquivo, estimar_linhas_xlsx,
                             ler_em_blocos, ler_planilhas)
from app.utils.erros import ErroValidacaoArquivo


@pytest.fixture
def limite_baixo(monkeypatch):
    """Simula um limite baixo (sem precisar gerar dezenas de milhares de
    linhas nos testes) para exercitar a trava de forma rápida."""
    monkeypatch.setattr(config, "LIMITE_LINHAS_ARQUIVO", 5)
    return 5


def test_arquivo_dentro_do_limite_e_lido_normalmente(planilhas):
    planilhas_lidas = ler_planilhas(planilhas["vendas"])
    assert planilhas_lidas
    assert not planilhas_lidas[0].dados.empty


def test_xlsx_acima_do_limite_e_recusado_sem_carregar_tudo(planilhas, limite_baixo):
    with pytest.raises(ErroValidacaoArquivo) as excinfo:
        ler_planilhas(planilhas["vendas"])
    mensagem = str(excinfo.value)
    assert "linhas" in mensagem
    assert "5" in mensagem
    assert any("Divida o arquivo" in d for d in excinfo.value.detalhes)


def test_csv_acima_do_limite_e_recusado(tmp_path, limite_baixo):
    dados = pd.DataFrame([
        {"Data": "01/08/2026", "Cidade": "Rio Bonito", "Equipe": "Equipe 01",
         "Canal": "Comercial", "Matrícula": i, "Quantidade": 1, "Valor": 100.0}
        for i in range(20)
    ])
    caminho = tmp_path / "venda_grande.csv"
    dados.to_csv(caminho, index=False, sep=";")

    with pytest.raises(ErroValidacaoArquivo) as excinfo:
        ler_planilhas(caminho)
    assert "linhas" in str(excinfo.value)


def test_mensagem_do_limite_sugere_dividir_ou_migrar_de_plano(planilhas, limite_baixo):
    with pytest.raises(ErroValidacaoArquivo) as excinfo:
        ler_planilhas(planilhas["vendas"])
    detalhes = " ".join(excinfo.value.detalhes)
    assert "Divida o arquivo" in detalhes
    assert "plano com mais memória" in detalhes


def test_dimensao_fantasma_nao_e_recusada_por_engano(tmp_path):
    """Regressão: uma planilha com só 20 linhas de dado real, mas que teve
    uma célula bem distante "tocada" (formatação arrastada, comum em
    exportações de outros sistemas), faz o `max_row` do openpyxl mentir
    (reporta a linha distante como se fosse o fim dos dados). Sem a
    confirmação por varredura real, isso recusaria por engano um arquivo
    pequeno de verdade."""
    caminho = tmp_path / "fantasma.xlsx"
    livro = openpyxl.Workbook()
    aba = livro.active
    aba.cell(row=1, column=1, value="Coluna A")  # cabeçalho, todo texto
    aba.cell(row=1, column=2, value="Coluna B")
    for linha in range(2, 22):  # 20 linhas de dado real
        aba.cell(row=linha, column=1, value=f"linha{linha}")
        aba.cell(row=linha, column=2, value=linha)
    aba.cell(row=300_000, column=5)  # célula "tocada" sem valor real
    livro.save(caminho)

    assert estimar_linhas_xlsx(caminho, limite=100) == 21  # cabeçalho + 20 linhas de dado
    planilhas_lidas = ler_planilhas(caminho)  # não deve levantar ErroValidacaoArquivo
    assert len(planilhas_lidas[0].dados) == 20


def test_dimensao_fantasma_muito_grande_ainda_e_recusada(tmp_path, limite_baixo):
    """A confirmação por varredura tem um teto (`LIMITE_VARREDURA_FANTASMA`)
    — um arquivo genuinamente grande continua sendo recusado, só que a
    confirmação para assim que tem certeza (ultrapassou o limite), sem
    varrer o arquivo inteiro."""
    caminho = tmp_path / "grande_de_verdade.xlsx"
    livro = openpyxl.Workbook()
    aba = livro.active
    for linha in range(1, 51):  # acima do limite_baixo (5), com dado real
        aba.cell(row=linha, column=1, value=f"linha{linha}")
    livro.save(caminho)

    with pytest.raises(ErroValidacaoArquivo):
        ler_planilhas(caminho)


def test_estimar_linhas_arquivo_despacha_por_extensao(planilhas, tmp_path):
    assert estimar_linhas_arquivo(planilhas["vendas"]) > 0

    csv = tmp_path / "a.csv"
    csv.write_text("a;b\n1;2\n3;4\n")
    assert estimar_linhas_arquivo(csv) == 3

    xls_desconhecido = tmp_path / "legado.xls"
    xls_desconhecido.write_bytes(b"nao e um xls de verdade")
    assert estimar_linhas_arquivo(xls_desconhecido) == 0


def _planilha_grande(caminho, linhas: int):
    """Arquivo com cabeçalho de Venda e `linhas` registros válidos."""
    livro = openpyxl.Workbook()
    aba = livro.active
    for coluna, nome in enumerate(
            ["Data", "Cidade", "Equipe", "Canal", "Matrícula", "Quantidade", "Valor"], start=1):
        aba.cell(row=1, column=coluna, value=nome)
    for i in range(linhas):
        linha = i + 2
        aba.cell(row=linha, column=1, value="01/08/2026")
        aba.cell(row=linha, column=2, value="Rio Bonito")
        aba.cell(row=linha, column=3, value=f"Equipe {i % 5:02d}")
        aba.cell(row=linha, column=4, value="Comercial")
        aba.cell(row=linha, column=5, value=f"M{i}")
        aba.cell(row=linha, column=6, value=1)
        aba.cell(row=linha, column=7, value=10.0)
    livro.save(caminho)


def test_ler_em_blocos_cobre_todas_as_linhas_sem_perder_nem_repetir(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "LINHAS_POR_BLOCO", 400)
    caminho = tmp_path / "grande.xlsx"
    _planilha_grande(caminho, 1000)

    blocos = list(ler_em_blocos(caminho))
    assert len(blocos) > 1, "deveria ter sido dividido em vários blocos"
    assert sum(len(b.dados) for b in blocos) == 1000

    # O índice preservado é o que permite apontar a linha real do Excel.
    indices = [i for b in blocos for i in b.dados.index]
    assert len(set(indices)) == len(indices), "índices repetidos entre blocos"
    assert min(indices) == 1 and max(indices) == 1000

    # Todos os blocos enxergam o mesmo cabeçalho.
    for bloco in blocos:
        assert "Matrícula" in list(bloco.dados.columns)


def test_arquivo_acima_do_limite_e_processado_em_blocos_e_nao_recusado(tmp_path, monkeypatch):
    """Regressão: antes, um arquivo grande era recusado. Agora ele deve
    carregar por completo, lido em pedaços."""
    from app.etl.pipeline import processar_arquivo
    from app.models.db import sessao

    monkeypatch.setattr(config, "LIMITE_LINHAS_ARQUIVO", 300)
    monkeypatch.setattr(config, "LINHAS_POR_BLOCO", 200)
    caminho = tmp_path / "venda_grande.xlsx"
    _planilha_grande(caminho, 900)

    with sessao() as s:
        resultado = processar_arquivo(s, caminho, dataset_forcado="vendas", arquivar=False)

    assert resultado.status in ("SUCESSO", "ATENCAO"), resultado.mensagem
    assert resultado.lidos == 900
    assert resultado.inseridos == 900
